from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import csv
import os
import re
import hashlib
from datetime import datetime, timedelta
import json
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import requests
import xml.etree.ElementTree as ET
from io import BytesIO
import PyPDF2
import logging
from functools import wraps

app = Flask(__name__)
app.config['SECRET_KEY'] = 'sua_chave_secreta_aqui'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///conciliacao.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Configurações de API bancária (exemplo)
app.config['BANK_API_KEY'] = os.environ.get('BANK_API_KEY', '')
app.config['BANK_API_URL'] = os.environ.get('BANK_API_URL', '')

# Configurações de ERP
app.config['ERP_API_KEY'] = os.environ.get('ERP_API_KEY', '')
app.config['ERP_API_URL'] = os.environ.get('ERP_API_URL', '')

# Criar pastas necessárias
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('logs', exist_ok=True)

# Configuração de logging
logging.basicConfig(
    filename='logs/conciliacao.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

db = SQLAlchemy(app)
CORS(app)

# Configuração do Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Verificação de usuário ativo antes de cada requisição
@app.before_request
def verificar_usuario_ativo():
    """Verifica se o usuário logado ainda está ativo"""
    # Excluir rotas que não precisam de verificação
    rotas_excluidas = ['login', 'logout', 'static']
    
    # Verificar se a rota atual está nas excluídas
    if request.endpoint in rotas_excluidas:
        return
    
    # Se o usuário está logado, verificar se ainda está ativo
    if current_user.is_authenticated:
        # Recarregar o usuário do banco para pegar o status atual
        usuario_atual = Usuario.query.get(current_user.id)
        if not usuario_atual or not usuario_atual.ativo:
            logout_user()
            flash('Sua conta foi desativada. Entre em contato com o administrador.', 'error')
            return redirect(url_for('login'))

# Funções de controle de acesso
def requires_permission(action='read'):
    """Decorator para controlar permissões de acesso"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({'error': 'Usuário não autenticado'}), 401
            
            # Admins têm acesso total
            if current_user.perfil == 'admin':
                return f(*args, **kwargs)
            
            # Auditores podem ler tudo, mas não modificar
            if current_user.perfil == 'auditor' and action == 'read':
                return f(*args, **kwargs)
            elif current_user.perfil == 'auditor' and action != 'read':
                return jsonify({'error': 'Auditores não podem modificar dados'}), 403
            
            # Usuários normais só acessam seus próprios dados
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def filter_by_user_access(query, model, user=None):
    """Filtra query baseado no acesso do usuário"""
    if not user:
        user = current_user
    
    # Admin e Auditor veem tudo
    if user.perfil in ['admin', 'auditor']:
        return query
    
    # Usuário normal vê apenas seus dados
    if hasattr(model, 'usuario_id'):
        return query.filter(model.usuario_id == user.id)
    
    return query

# Modelos do banco de dados
class Usuario(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nome_completo = db.Column(db.String(200), nullable=False)
    perfil = db.Column(db.String(50), default='usuario')  # admin, usuario, auditor
    ativo = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acesso = db.Column(db.DateTime)

class ContaBancaria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))  # Controle de acesso
    banco = db.Column(db.String(100), nullable=False)
    agencia = db.Column(db.String(20), nullable=False)
    conta = db.Column(db.String(20), nullable=False)
    tipo_conta = db.Column(db.String(50), default='corrente')  # corrente, poupanca, investimento
    saldo_atual = db.Column(db.Float, default=0.0)
    ativa = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    usuario = db.relationship('Usuario', backref='contas_bancarias')

class ExtratoBancario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    conta_id = db.Column(db.Integer, db.ForeignKey('conta_bancaria.id'))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))  # Controle de acesso
    data = db.Column(db.Date, nullable=False)
    descricao = db.Column(db.String(500), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # 'credito' ou 'debito'
    categoria = db.Column(db.String(100), default='Não categorizado')
    conciliado = db.Column(db.Boolean, default=False)
    arquivo_origem = db.Column(db.String(200))
    formato_arquivo = db.Column(db.String(20))  # CSV, OFX, CNAB, PDF, API
    numero_documento = db.Column(db.String(100))
    hash_transacao = db.Column(db.String(64))  # Para detectar duplicatas
    transacao_recorrente = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    conta = db.relationship('ContaBancaria', backref='extratos')
    usuario = db.relationship('Usuario', backref='extratos_bancarios')

class LancamentoContabil(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))  # Controle de acesso
    data = db.Column(db.Date, nullable=False)
    descricao = db.Column(db.String(500), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # 'credito' ou 'debito'
    categoria = db.Column(db.String(100), default='Não categorizado')
    conciliado = db.Column(db.Boolean, default=False)
    arquivo_origem = db.Column(db.String(200))
    numero_documento = db.Column(db.String(100))
    centro_custo = db.Column(db.String(100))
    conta_contabil = db.Column(db.String(100))
    fornecedor_cliente = db.Column(db.String(200))
    hash_transacao = db.Column(db.String(64))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    usuario = db.relationship('Usuario', backref='lancamentos_contabeis')

class Conciliacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    extrato_id = db.Column(db.Integer, db.ForeignKey('extrato_bancario.id'))
    lancamento_id = db.Column(db.Integer, db.ForeignKey('lancamento_contabil.id'))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    data_conciliacao = db.Column(db.DateTime, default=datetime.utcnow)
    observacoes = db.Column(db.Text)
    tipo_conciliacao = db.Column(db.String(20), default='manual')  # manual, automatica
    status = db.Column(db.String(20), default='ativa')  # ativa, cancelada
    
    extrato = db.relationship('ExtratoBancario', backref='conciliacoes')
    lancamento = db.relationship('LancamentoContabil', backref='conciliacoes')
    usuario = db.relationship('Usuario', backref='conciliacoes')

class RegraConciliacao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.String(500))
    tipo_regra = db.Column(db.String(50), nullable=False)  # valor, descricao, documento, recorrente
    valor_exato = db.Column(db.Float)
    valor_min = db.Column(db.Float)
    valor_max = db.Column(db.Float)
    descricao_padrao = db.Column(db.String(500))
    numero_documento = db.Column(db.String(100))
    categoria_destino = db.Column(db.String(100))
    ativa = db.Column(db.Boolean, default=True)
    prioridade = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Divergencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), nullable=False)  # duplicata, valor_incorreto, nao_reconhecido
    extrato_id = db.Column(db.Integer, db.ForeignKey('extrato_bancario.id'))
    lancamento_id = db.Column(db.Integer, db.ForeignKey('lancamento_contabil.id'))
    descricao = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='pendente')  # pendente, resolvida, ignorada
    justificativa = db.Column(db.Text)
    usuario_resolucao = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    data_resolucao = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    extrato = db.relationship('ExtratoBancario', backref='divergencias')
    lancamento = db.relationship('LancamentoContabil', backref='divergencias')
    usuario = db.relationship('Usuario', backref='divergencias_resolvidas')

class LogAuditoria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    acao = db.Column(db.String(100), nullable=False)
    tabela = db.Column(db.String(100))
    registro_id = db.Column(db.Integer)
    dados_anteriores = db.Column(db.Text)
    dados_novos = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    usuario = db.relationship('Usuario', backref='logs_auditoria')

class ConfiguracaoAPI(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    tipo = db.Column(db.String(50), nullable=False)  # banco, erp
    url = db.Column(db.String(500), nullable=False)
    api_key = db.Column(db.String(200))
    usuario = db.Column(db.String(100))
    senha_hash = db.Column(db.String(200))
    ativa = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Criar tabelas
with app.app_context():
    db.create_all()
    
    # Criar usuário admin padrão se não existir
    admin = Usuario.query.filter_by(username='admin').first()
    if not admin:
        admin = Usuario(
            username='admin',
            email='admin@conciliacao.com',
            password_hash=generate_password_hash('admin123'),
            nome_completo='Administrador do Sistema',
            perfil='admin'
        )
        db.session.add(admin)
        db.session.commit()

@login_manager.user_loader
def load_user(user_id):
    usuario = Usuario.query.get(int(user_id))
    # Retornar None se o usuário não existir ou estiver inativo
    if usuario and usuario.ativo:
        return usuario
    return None

def log_auditoria(acao, tabela=None, registro_id=None, dados_anteriores=None, dados_novos=None):
    """Registra ação na auditoria"""
    try:
        log = LogAuditoria(
            usuario_id=current_user.id if current_user.is_authenticated else None,
            acao=acao,
            tabela=tabela,
            registro_id=registro_id,
            dados_anteriores=json.dumps(dados_anteriores) if dados_anteriores else None,
            dados_novos=json.dumps(dados_novos) if dados_novos else None,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logging.error(f"Erro ao registrar log de auditoria: {e}")

def admin_required(f):
    """Decorator para verificar se o usuário é admin"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.perfil != 'admin':
            return jsonify({'error': 'Acesso negado'}), 403
        return f(*args, **kwargs)
    return decorated_function

def parse_date(date_str):
    """Converte string de data para objeto date"""
    try:
        # Tentar diferentes formatos de data
        formats = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']
        for fmt in formats:
            try:
                return datetime.strptime(str(date_str).strip(), fmt).date()
            except ValueError:
                continue
        # Se nenhum formato funcionar, retornar data atual
        return datetime.now().date()
    except:
        return datetime.now().date()

def parse_value(value_str):
    """Converte string de valor para float"""
    try:
        # Remover caracteres não numéricos exceto ponto e vírgula
        value_str = str(value_str).replace('R$', '').replace('$', '').replace(' ', '').replace(',', '.')
        return float(value_str)
    except:
        return 0.0

def generate_hash(descricao, valor, data, tipo):
    """Gera hash único para transação"""
    hash_string = f"{descricao}{valor}{data}{tipo}"
    return hashlib.sha256(hash_string.encode()).hexdigest()

def detect_duplicate_transaction(descricao, valor, data, tipo, tabela='extrato'):
    """Detecta transações duplicadas"""
    hash_transacao = generate_hash(descricao, valor, data, tipo)
    
    if tabela == 'extrato':
        return ExtratoBancario.query.filter_by(hash_transacao=hash_transacao).first()
    else:
        return LancamentoContabil.query.filter_by(hash_transacao=hash_transacao).first()

def process_ofx_file(filepath):
    """Processa arquivo OFX"""
    registros = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            content = file.read()
            
        # Parse OFX (simplificado)
        # Em produção, usar biblioteca específica como ofxparse
        transactions = re.findall(r'<STMTTRN>.*?<TRNAMT>(.*?)</TRNAMT>.*?<FITID>(.*?)</FITID>.*?<MEMO>(.*?)</MEMO>.*?<DTPOSTED>(.*?)</DTPOSTED>.*?</STMTTRN>', content, re.DOTALL)
        
        for trn in transactions:
            try:
                valor = parse_value(trn[0])
                descricao = trn[2].strip()
                data_str = trn[3][:8]  # YYYYMMDD
                data = datetime.strptime(data_str, '%Y%m%d').date()
                tipo = 'credito' if valor > 0 else 'debito'
                
                registros.append({
                    'data': data,
                    'descricao': descricao,
                    'valor': valor,
                    'tipo': tipo,
                    'numero_documento': trn[1]
                })
            except Exception as e:
                continue
                
    except Exception as e:
        logging.error(f"Erro ao processar OFX: {e}")
    
    return registros

def process_cnab_file(filepath):
    """Processa arquivo CNAB"""
    registros = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            for line in file:
                if len(line) < 240:  # CNAB padrão
                    continue
                    
                try:
                    # Parse CNAB (simplificado)
                    data_str = line[6:14]  # Data
                    valor_str = line[119:134]  # Valor
                    descricao = line[43:73].strip()  # Histórico
                    
                    data = datetime.strptime(data_str, '%d%m%Y').date()
                    valor = parse_value(valor_str) / 100  # CNAB usa centavos
                    tipo = 'credito' if valor > 0 else 'debito'
                    
                    registros.append({
                        'data': data,
                        'descricao': descricao,
                        'valor': valor,
                        'tipo': tipo
                    })
                except Exception as e:
                    continue
                    
    except Exception as e:
        logging.error(f"Erro ao processar CNAB: {e}")
    
    return registros

def process_pdf_file(filepath):
    """Processa arquivo PDF"""
    registros = []
    try:
        with open(filepath, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            
            for page in pdf_reader.pages:
                text += page.extract_text()
            
            # Parse texto do PDF (simplificado)
            # Em produção, usar regex mais sofisticado
            lines = text.split('\n')
            for line in lines:
                # Procurar padrões de data, valor e descrição
                date_pattern = r'(\d{2}/\d{2}/\d{4})'
                value_pattern = r'R?\$?\s*([\d.,]+)'
                
                date_match = re.search(date_pattern, line)
                value_match = re.search(value_pattern, line)
                
                if date_match and value_match:
                    try:
                        data = parse_date(date_match.group(1))
                        valor = parse_value(value_match.group(1))
                        descricao = line.replace(date_match.group(0), '').replace(value_match.group(0), '').strip()
                        
                        if descricao and valor != 0:
                            tipo = 'credito' if valor > 0 else 'debito'
                            registros.append({
                                'data': data,
                                'descricao': descricao,
                                'valor': valor,
                                'tipo': tipo
                            })
                    except Exception as e:
                        continue
                        
    except Exception as e:
        logging.error(f"Erro ao processar PDF: {e}")
    
    return registros

def fetch_bank_api_data(conta_id, data_inicio, data_fim):
    """Busca dados da API bancária"""
    registros = []
    try:
        config = ConfiguracaoAPI.query.filter_by(tipo='banco', ativa=True).first()
        if not config:
            return registros
            
        headers = {
            'Authorization': f'Bearer {config.api_key}',
            'Content-Type': 'application/json'
        }
        
        params = {
            'conta_id': conta_id,
            'data_inicio': data_inicio.strftime('%Y-%m-%d'),
            'data_fim': data_fim.strftime('%Y-%m-%d')
        }
        
        response = requests.get(f"{config.url}/transacoes", headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            for transacao in data.get('transacoes', []):
                registros.append({
                    'data': parse_date(transacao['data']),
                    'descricao': transacao['descricao'],
                    'valor': parse_value(transacao['valor']),
                    'tipo': transacao['tipo'],
                    'numero_documento': transacao.get('documento', '')
                })
                
    except Exception as e:
        logging.error(f"Erro ao buscar dados da API bancária: {e}")
    
    return registros

def fetch_erp_data(data_inicio, data_fim):
    """Busca dados do ERP"""
    registros = []
    try:
        config = ConfiguracaoAPI.query.filter_by(tipo='erp', ativa=True).first()
        if not config:
            return registros
            
        headers = {
            'Authorization': f'Bearer {config.api_key}',
            'Content-Type': 'application/json'
        }
        
        params = {
            'data_inicio': data_inicio.strftime('%Y-%m-%d'),
            'data_fim': data_fim.strftime('%Y-%m-%d')
        }
        
        response = requests.get(f"{config.url}/lancamentos", headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            for lancamento in data.get('lancamentos', []):
                registros.append({
                    'data': parse_date(lancamento['data']),
                    'descricao': lancamento['descricao'],
                    'valor': parse_value(lancamento['valor']),
                    'tipo': lancamento['tipo'],
                    'numero_documento': lancamento.get('documento', ''),
                    'centro_custo': lancamento.get('centro_custo', ''),
                    'conta_contabil': lancamento.get('conta_contabil', ''),
                    'fornecedor_cliente': lancamento.get('fornecedor_cliente', '')
                })
                
    except Exception as e:
        logging.error(f"Erro ao buscar dados do ERP: {e}")
    
    return registros

def detect_columns(headers):
    """Detecta automaticamente as colunas baseado nos nomes"""
    colunas_encontradas = []
    
    for i, header in enumerate(headers):
        header_lower = str(header).lower()
        if 'data' in header_lower or 'date' in header_lower:
            colunas_encontradas.append(('data', i))
        elif 'desc' in header_lower or 'historico' in header_lower or 'descrição' in header_lower:
            colunas_encontradas.append(('descricao', i))
        elif 'valor' in header_lower or 'amount' in header_lower:
            colunas_encontradas.append(('valor', i))
        elif 'tipo' in header_lower or 'type' in header_lower:
            colunas_encontradas.append(('tipo', i))
    
    # Se não encontrou colunas, usar as primeiras
    if len(colunas_encontradas) < 3:
        colunas_encontradas = [
            ('data', 0),
            ('descricao', 1),
            ('valor', 2),
            ('tipo', 3) if len(headers) > 3 else ('tipo', None)
        ]
    
    return colunas_encontradas

def process_csv_file(filepath):
    """Processa arquivo CSV"""
    registros = []
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            headers = next(reader)  # Primeira linha são os cabeçalhos
            colunas = detect_columns(headers)
            
            for row in reader:
                if len(row) < 3:
                    continue
                
                try:
                    data = parse_date(row[colunas[0][1]])
                    descricao = str(row[colunas[1][1]])
                    valor = parse_value(row[colunas[2][1]])
                    
                    # Determinar tipo baseado no valor ou coluna
                    if len(colunas) > 3 and colunas[3][1] is not None:
                        tipo = str(row[colunas[3][1]]).lower()
                    else:
                        tipo = 'credito' if valor > 0 else 'debito'
                    
                    if tipo not in ['credito', 'debito']:
                        tipo = 'credito' if valor > 0 else 'debito'
                    
                    registros.append({
                        'data': data,
                        'descricao': descricao,
                        'valor': valor,
                        'tipo': tipo
                    })
                except Exception as e:
                    continue
    except Exception as e:
        print(f"Erro ao processar CSV: {e}")
    
    return registros

def process_excel_file(filepath):
    """Processa arquivo Excel"""
    registros = []
    try:
        workbook = load_workbook(filepath, read_only=True)
        worksheet = workbook.active
        
        # Ler cabeçalhos
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)
        
        colunas = detect_columns(headers)
        
        # Ler dados
        for row in worksheet.iter_rows(min_row=2):
            if len(row) < 3:
                continue
            
            try:
                data = parse_date(row[colunas[0][1]].value)
                descricao = str(row[colunas[1][1]].value or '')
                valor = parse_value(row[colunas[2][1]].value)
                
                # Determinar tipo baseado no valor ou coluna
                if len(colunas) > 3 and colunas[3][1] is not None:
                    tipo = str(row[colunas[3][1]].value or '').lower()
                else:
                    tipo = 'credito' if valor > 0 else 'debito'
                
                if tipo not in ['credito', 'debito']:
                    tipo = 'credito' if valor > 0 else 'debito'
                
                registros.append({
                    'data': data,
                    'descricao': descricao,
                    'valor': valor,
                    'tipo': tipo
                })
            except Exception as e:
                continue
                
    except Exception as e:
        print(f"Erro ao processar Excel: {e}")
    
    return registros

def aplicar_regras_conciliacao(registro, tipo='extrato'):
    """Aplica regras de conciliação automática"""
    regras = RegraConciliacao.query.filter_by(ativa=True).order_by(RegraConciliacao.prioridade.desc()).all()
    
    for regra in regras:
        if regra.tipo_regra == 'valor':
            if regra.valor_exato and abs(registro['valor'] - regra.valor_exato) < 0.01:
                registro['categoria'] = regra.categoria_destino
                return True
            elif regra.valor_min and regra.valor_max:
                if regra.valor_min <= registro['valor'] <= regra.valor_max:
                    registro['categoria'] = regra.categoria_destino
                    return True
        elif regra.tipo_regra == 'descricao':
            if regra.descricao_padrao and regra.descricao_padrao.lower() in registro['descricao'].lower():
                registro['categoria'] = regra.categoria_destino
                return True
        elif regra.tipo_regra == 'documento':
            if regra.numero_documento and regra.numero_documento in registro.get('numero_documento', ''):
                registro['categoria'] = regra.categoria_destino
                return True
    
    return False

def detectar_transacoes_recorrentes():
    """Detecta transações recorrentes baseado em padrões"""
    extratos = ExtratoBancario.query.filter_by(transacao_recorrente=False).all()
    
    for extrato in extratos:
        # Buscar transações similares
        similares = ExtratoBancario.query.filter(
            ExtratoBancario.descricao == extrato.descricao,
            ExtratoBancario.valor == extrato.valor,
            ExtratoBancario.id != extrato.id
        ).all()
        
        if len(similares) >= 2:  # Se encontrou pelo menos 2 transações similares
            extrato.transacao_recorrente = True
            for similar in similares:
                similar.transacao_recorrente = True
    
    db.session.commit()

def conciliacao_automatica():
    """Realiza conciliação automática baseada em regras"""
    extratos_nao_conciliados = ExtratoBancario.query.filter_by(conciliado=False).all()
    lancamentos_nao_conciliados = LancamentoContabil.query.filter_by(conciliado=False).all()
    
    conciliacoes_realizadas = 0
    
    for extrato in extratos_nao_conciliados:
        for lancamento in lancamentos_nao_conciliados:
            # Verificar se já estão conciliados
            if extrato.conciliado or lancamento.conciliado:
                continue
                
            # Regras de conciliação automática
            if (extrato.valor == lancamento.valor and 
                extrato.data == lancamento.data and
                extrato.tipo == lancamento.tipo):
                
                # Verificar se já existe conciliação
                conciliacao_existente = Conciliacao.query.filter_by(
                    extrato_id=extrato.id,
                    lancamento_id=lancamento.id
                ).first()
                
                if not conciliacao_existente:
                    # Criar conciliação automática
                    conciliacao = Conciliacao(
                        extrato_id=extrato.id,
                        lancamento_id=lancamento.id,
                        usuario_id=1,  # Sistema
                        tipo_conciliacao='automatica',
                        observacoes='Conciliação automática por valor, data e tipo'
                    )
                    db.session.add(conciliacao)
                    
                    # Marcar como conciliado
                    extrato.conciliado = True
                    lancamento.conciliado = True
                    
                    conciliacoes_realizadas += 1
                    
                    # Registrar na auditoria
                    log_auditoria(
                        'conciliacao_automatica',
                        'conciliacao',
                        conciliacao.id,
                        None,
                        {'extrato_id': extrato.id, 'lancamento_id': lancamento.id}
                    )
    
    db.session.commit()
    return conciliacoes_realizadas

def verificar_divergencias():
    """Verifica e cria alertas de divergências"""
    # Verificar duplicatas
    extratos = ExtratoBancario.query.all()
    for extrato in extratos:
        duplicatas = ExtratoBancario.query.filter(
            ExtratoBancario.descricao == extrato.descricao,
            ExtratoBancario.valor == extrato.valor,
            ExtratoBancario.data == extrato.data,
            ExtratoBancario.id != extrato.id
        ).all()
        
        if duplicatas:
            for duplicata in duplicatas:
                divergencia = Divergencia(
                    tipo='duplicata',
                    extrato_id=extrato.id,
                    descricao=f'Transação duplicada detectada: {extrato.descricao} - R$ {extrato.valor}'
                )
                db.session.add(divergencia)
    
    # Verificar valores incorretos (diferença > 5%)
    conciliacoes = Conciliacao.query.filter_by(status='ativa').all()
    for conciliacao in conciliacoes:
        if abs(conciliacao.extrato.valor - conciliacao.lancamento.valor) > 0.01:
            divergencia = Divergencia(
                tipo='valor_incorreto',
                extrato_id=conciliacao.extrato_id,
                lancamento_id=conciliacao.lancamento_id,
                descricao=f'Diferença de valor detectada: Extrato R$ {conciliacao.extrato.valor} vs Lançamento R$ {conciliacao.lancamento.valor}'
            )
            db.session.add(divergencia)
    
    db.session.commit()

@app.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect(url_for('register'))
    
    # Dados do usuário para passar ao template
    user_data = {
        'nome': current_user.nome_completo,
        'email': current_user.email,
        'username': current_user.username,
        'perfil': current_user.perfil,
        'ultimo_acesso': current_user.ultimo_acesso.strftime('%d/%m/%Y %H:%M') if current_user.ultimo_acesso else 'Nunca'
    }
    
    return render_template('index.html', user=user_data)

@app.route('/admin')
@login_required
@admin_required
def admin_panel():
    """Painel de administração - apenas para administradores"""
    return render_template('admin.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')
    return render_template('register.html')

@app.route('/api/register', methods=['POST'])
def api_register():
    try:
        data = request.get_json()
        
        # Validar dados necessários
        required_fields = ['username', 'email', 'password', 'nome_completo']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'O campo {field} é obrigatório'}), 400
        
        # Verificar se usuário já existe
        if Usuario.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'Nome de usuário já existe'}), 400
        
        # Verificar se email já existe
        if Usuario.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email já está em uso'}), 400
        
        # Criar novo usuário
        novo_usuario = Usuario(
            username=data['username'],
            email=data['email'],
            password_hash=generate_password_hash(data['password']),
            nome_completo=data['nome_completo'],
            perfil='usuario'  # Perfil padrão para novos usuários
        )
        
        db.session.add(novo_usuario)
        db.session.commit()
        
        log_auditoria(
            'criar_usuario',
            'usuario',
            novo_usuario.id,
            None,
            {'username': data['username'], 'email': data['email']}
        )
        
        return jsonify({
            'success': True,
            'message': 'Usuário criado com sucesso'
        })
        
    except Exception as e:
        logging.error(f"Erro ao criar usuário: {e}")
        return jsonify({'error': 'Erro ao criar usuário'}), 500

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Tentar primeiro obter dados JSON (para requisições AJAX)
        if request.is_json:
            data = request.get_json()
            username = data.get('username')
            password = data.get('password')
        else:
            # Fallback para dados de formulário HTML
            username = request.form.get('username')
            password = request.form.get('password')
        
        if not username or not password:
            if request.is_json:
                return jsonify({'error': 'Usuário e senha são obrigatórios'}), 400
            else:
                flash('Usuário e senha são obrigatórios', 'error')
                return render_template('login.html')
        
        user = Usuario.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            # Verificar se o usuário está ativo
            if not user.ativo:
                if request.is_json:
                    return jsonify({'error': 'Usuário bloqueado. Entre em contato com o administrador.'}), 403
                else:
                    flash('Usuário bloqueado. Entre em contato com o administrador.', 'error')
                    return render_template('login.html')
            
            login_user(user)
            user.ultimo_acesso = datetime.utcnow()
            db.session.commit()
            
            log_auditoria('login', 'usuario', user.id)
            
            if request.is_json:
                return jsonify({'success': True, 'redirect': url_for('index')})
            else:
                return redirect(url_for('index'))
        else:
            if request.is_json:
                return jsonify({'error': 'Usuário ou senha inválidos'}), 401
            else:
                flash('Usuário ou senha inválidos', 'error')
                return render_template('login.html')
    
    return render_template('login.html')

@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    try:
        log_auditoria('logout', 'usuario', current_user.id)
        logout_user()
        
        # Se for uma requisição POST (AJAX), retornar JSON
        if request.method == 'POST':
            return jsonify({'success': True, 'message': 'Logout realizado com sucesso'})
        
        # Se for GET, redirecionar normalmente
        return redirect(url_for('login'))
    except Exception as e:
        logging.error(f"Erro no logout: {e}")
        if request.method == 'POST':
            return jsonify({'success': False, 'error': str(e)}), 500
        return redirect(url_for('login'))

# === ROTAS DE ADMINISTRAÇÃO ===

@app.route('/api/contas-bancarias', methods=['GET'])
@login_required
def get_contas_bancarias():
    contas = ContaBancaria.query.filter_by(ativa=True).all()
    return jsonify([{
        'id': c.id,
        'banco': c.banco,
        'agencia': c.agencia,
        'conta': c.conta,
        'tipo_conta': c.tipo_conta,
        'saldo_atual': c.saldo_atual
    } for c in contas])

@app.route('/api/contas-bancarias', methods=['POST'])
@login_required
@admin_required
def criar_conta_bancaria():
    data = request.get_json()
    
    conta = ContaBancaria(
        banco=data['banco'],
        agencia=data['agencia'],
        conta=data['conta'],
        tipo_conta=data.get('tipo_conta', 'corrente'),
        saldo_atual=data.get('saldo_atual', 0.0)
    )
    
    db.session.add(conta)
    db.session.commit()
    
    log_auditoria('criar_conta_bancaria', 'conta_bancaria', conta.id, None, data)
    return jsonify({'success': True, 'message': 'Conta bancária criada com sucesso'})

@app.route('/api/regras-conciliacao', methods=['GET'])
@login_required
def get_regras_conciliacao():
    regras = RegraConciliacao.query.filter_by(ativa=True).order_by(RegraConciliacao.prioridade.desc()).all()
    return jsonify([{
        'id': r.id,
        'nome': r.nome,
        'descricao': r.descricao,
        'tipo_regra': r.tipo_regra,
        'valor_exato': r.valor_exato,
        'valor_min': r.valor_min,
        'valor_max': r.valor_max,
        'descricao_padrao': r.descricao_padrao,
        'numero_documento': r.numero_documento,
        'categoria_destino': r.categoria_destino,
        'prioridade': r.prioridade
    } for r in regras])

@app.route('/api/regras-conciliacao', methods=['POST'])
@login_required
@admin_required
def criar_regra_conciliacao():
    data = request.get_json()
    
    regra = RegraConciliacao(
        nome=data['nome'],
        descricao=data.get('descricao', ''),
        tipo_regra=data['tipo_regra'],
        valor_exato=data.get('valor_exato'),
        valor_min=data.get('valor_min'),
        valor_max=data.get('valor_max'),
        descricao_padrao=data.get('descricao_padrao'),
        numero_documento=data.get('numero_documento'),
        categoria_destino=data['categoria_destino'],
        prioridade=data.get('prioridade', 1)
    )
    
    db.session.add(regra)
    db.session.commit()
    
    log_auditoria('criar_regra_conciliacao', 'regra_conciliacao', regra.id, None, data)
    return jsonify({'success': True, 'message': 'Regra criada com sucesso'})

@app.route('/api/conciliacao-automatica', methods=['POST'])
@login_required
def executar_conciliacao_automatica():
    try:
        # Detectar transações recorrentes
        detectar_transacoes_recorrentes()
        
        # Executar conciliação automática
        conciliacoes = conciliacao_automatica()
        
        # Verificar divergências
        verificar_divergencias()
        
        log_auditoria('conciliacao_automatica', None, None, None, {'conciliacoes_realizadas': conciliacoes})
        
        return jsonify({
            'success': True,
            'message': f'Conciliação automática executada. {conciliacoes} conciliações realizadas.',
            'conciliados': conciliacoes
        })
    except Exception as e:
        logging.error(f"Erro na conciliação automática: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/conciliacoes', methods=['GET'])
@login_required
@requires_permission('read')
def get_conciliacoes():
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        tipo = request.args.get('tipo')
        status = request.args.get('status')
        usuario_id = request.args.get('usuario_id')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        
        # Construir query base com controle de acesso
        query = Conciliacao.query
        query = filter_by_user_access(query, Conciliacao)
        
        # Aplicar filtros
        if data_inicio:
            query = query.filter(Conciliacao.data_conciliacao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            query = query.filter(Conciliacao.data_conciliacao <= datetime.strptime(data_fim, '%Y-%m-%d'))
        if tipo:
            query = query.filter(Conciliacao.tipo_conciliacao == tipo)
        if status:
            query = query.filter(Conciliacao.status == status)
        
        # Apenas admin/auditor podem filtrar por outro usuário
        if usuario_id and current_user.perfil in ['admin', 'auditor']:
            query = query.filter(Conciliacao.usuario_id == usuario_id)
        
        # Ordenar por data mais recente
        query = query.order_by(Conciliacao.data_conciliacao.desc())
        
        # Paginação
        conciliacoes = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        # Formatear resposta
        resultado = {
            'conciliacoes': [{
                'id': c.id,
                'data_conciliacao': c.data_conciliacao.strftime('%Y-%m-%d %H:%M:%S'),
                'tipo_conciliacao': c.tipo_conciliacao,
                'status': c.status,
                'observacoes': c.observacoes,
                'usuario': {
                    'id': c.usuario.id,
                    'nome': c.usuario.nome_completo,
                    'username': c.usuario.username
                } if c.usuario else None,
                'extrato': {
                    'id': c.extrato.id,
                    'data': c.extrato.data.strftime('%Y-%m-%d'),
                    'descricao': c.extrato.descricao,
                    'valor': c.extrato.valor,
                    'tipo': c.extrato.tipo,
                    'conta': {
                        'banco': c.extrato.conta.banco,
                        'agencia': c.extrato.conta.agencia,
                        'conta': c.extrato.conta.conta
                    } if c.extrato.conta else None
                } if c.extrato else None,
                'lancamento': {
                    'id': c.lancamento.id,
                    'data': c.lancamento.data.strftime('%Y-%m-%d'),
                    'descricao': c.lancamento.descricao,
                    'valor': c.lancamento.valor,
                    'tipo': c.lancamento.tipo,
                    'categoria': c.lancamento.categoria
                } if c.lancamento else None
            } for c in conciliacoes.items],
            'pagination': {
                'page': conciliacoes.page,
                'pages': conciliacoes.pages,
                'per_page': conciliacoes.per_page,
                'total': conciliacoes.total,
                'has_next': conciliacoes.has_next,
                'has_prev': conciliacoes.has_prev
            }
        }
        
        return jsonify(resultado)
        
    except Exception as e:
        logging.error(f"Erro ao buscar conciliações: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/conciliacoes/estatisticas', methods=['GET'])
@login_required
def get_estatisticas_conciliacoes():
    try:
        # Estatísticas gerais
        total_conciliacoes = Conciliacao.query.count()
        conciliacoes_ativas = Conciliacao.query.filter_by(status='ativa').count()
        conciliacoes_canceladas = Conciliacao.query.filter_by(status='cancelada').count()
        conciliacoes_automaticas = Conciliacao.query.filter_by(tipo_conciliacao='automatica').count()
        conciliacoes_manuais = Conciliacao.query.filter_by(tipo_conciliacao='manual').count()
        
        # Estatísticas por período (últimos 30 dias)
        data_limite = datetime.utcnow() - timedelta(days=30)
        conciliacoes_recentes = Conciliacao.query.filter(
            Conciliacao.data_conciliacao >= data_limite
        ).count()
        
        # Top 5 usuários que mais conciliaram
        top_usuarios = db.session.query(
            Usuario.nome_completo,
            db.func.count(Conciliacao.id).label('total_conciliacoes')
        ).join(Conciliacao).group_by(Usuario.id).order_by(
            db.desc('total_conciliacoes')
        ).limit(5).all()
        
        # Conciliações por dia dos últimos 7 dias
        conciliacoes_por_dia = []
        for i in range(7):
            data = datetime.utcnow() - timedelta(days=i)
            count = Conciliacao.query.filter(
                db.func.date(Conciliacao.data_conciliacao) == data.date()
            ).count()
            conciliacoes_por_dia.append({
                'data': data.strftime('%Y-%m-%d'),
                'count': count
            })
        
        return jsonify({
            'total_conciliacoes': total_conciliacoes,
            'conciliacoes_ativas': conciliacoes_ativas,
            'conciliacoes_canceladas': conciliacoes_canceladas,
            'conciliacoes_automaticas': conciliacoes_automaticas,
            'conciliacoes_manuais': conciliacoes_manuais,
            'conciliacoes_recentes': conciliacoes_recentes,
            'top_usuarios': [{
                'nome': usuario[0],
                'total': usuario[1]
            } for usuario in top_usuarios],
            'conciliacoes_por_dia': conciliacoes_por_dia
        })
        
    except Exception as e:
        logging.error(f"Erro ao buscar estatísticas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/divergencias', methods=['GET'])
@login_required
def get_divergencias():
    divergencias = Divergencia.query.filter_by(status='pendente').order_by(Divergencia.created_at.desc()).all()
    return jsonify([{
        'id': d.id,
        'tipo': d.tipo,
        'descricao': d.descricao,
        'status': d.status,
        'created_at': d.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        'extrato': {
            'id': d.extrato.id,
            'descricao': d.extrato.descricao,
            'valor': d.extrato.valor,
            'data': d.extrato.data.strftime('%Y-%m-%d')
        } if d.extrato else None,
        'lancamento': {
            'id': d.lancamento.id,
            'descricao': d.lancamento.descricao,
            'valor': d.lancamento.valor,
            'data': d.lancamento.data.strftime('%Y-%m-%d')
        } if d.lancamento else None
    } for d in divergencias])

@app.route('/api/divergencias/<int:divergencia_id>/resolver', methods=['POST'])
@login_required
def resolver_divergencia(divergencia_id):
    data = request.get_json()
    divergencia = Divergencia.query.get(divergencia_id)
    
    if not divergencia:
        return jsonify({'error': 'Divergência não encontrada'}), 404
    
    divergencia.status = 'resolvida'
    divergencia.justificativa = data.get('justificativa', '')
    divergencia.usuario_resolucao = current_user.id
    divergencia.data_resolucao = datetime.utcnow()
    
    db.session.commit()
    
    log_auditoria('resolver_divergencia', 'divergencia', divergencia.id, None, data)
    return jsonify({'success': True, 'message': 'Divergência resolvida com sucesso'})

@app.route('/api/auditoria', methods=['GET'])
@login_required
@admin_required
def get_auditoria():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    acao_filtro = request.args.get('acao')
    data_filtro = request.args.get('data')
    
    query = LogAuditoria.query
    
    # Aplicar filtros
    if acao_filtro:
        query = query.filter(LogAuditoria.acao == acao_filtro)
    
    if data_filtro:
        try:
            data = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            query = query.filter(db.func.date(LogAuditoria.created_at) == data)
        except ValueError:
            pass
    
    logs = query.order_by(LogAuditoria.created_at.desc()).limit(per_page).all()
    
    return jsonify([{
        'id': l.id,
        'usuario': l.usuario.nome_completo if l.usuario else 'Sistema',
        'username': l.usuario.username if l.usuario else 'sistema',
        'acao': l.acao,
        'tabela': l.tabela,
        'registro_id': l.registro_id,
        'dados_anteriores': l.dados_anteriores,
        'dados_novos': l.dados_novos,
        'ip_address': l.ip_address,
        'user_agent': l.user_agent,
        'created_at': l.created_at.strftime('%Y-%m-%d %H:%M:%S')
    } for l in logs])

@app.route('/api/relatorios/consolidado', methods=['GET'])
@login_required
def relatorio_consolidado():
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    conta_id = request.args.get('conta_id')
    
    query_extratos = ExtratoBancario.query
    query_lancamentos = LancamentoContabil.query
    
    if data_inicio:
        query_extratos = query_extratos.filter(ExtratoBancario.data >= data_inicio)
        query_lancamentos = query_lancamentos.filter(LancamentoContabil.data >= data_inicio)
    
    if data_fim:
        query_extratos = query_extratos.filter(ExtratoBancario.data <= data_fim)
        query_lancamentos = query_lancamentos.filter(LancamentoContabil.data <= data_fim)
    
    if conta_id:
        query_extratos = query_extratos.filter(ExtratoBancario.conta_id == conta_id)
    
    extratos = query_extratos.all()
    lancamentos = query_lancamentos.all()
    
    # Estatísticas
    total_extratos = len(extratos)
    total_lancamentos = len(lancamentos)
    extratos_conciliados = len([e for e in extratos if e.conciliado])
    lancamentos_conciliados = len([l for l in lancamentos if l.conciliado])
    
    # Valores por categoria
    categorias_extratos = {}
    categorias_lancamentos = {}
    
    for extrato in extratos:
        cat = extrato.categoria
        if cat not in categorias_extratos:
            categorias_extratos[cat] = {'credito': 0, 'debito': 0}
        categorias_extratos[cat][extrato.tipo] += extrato.valor
    
    for lancamento in lancamentos:
        cat = lancamento.categoria
        if cat not in categorias_lancamentos:
            categorias_lancamentos[cat] = {'credito': 0, 'debito': 0}
        categorias_lancamentos[cat][lancamento.tipo] += lancamento.valor
    
    return jsonify({
        'periodo': {
            'inicio': data_inicio,
            'fim': data_fim
        },
        'estatisticas': {
            'total_extratos': total_extratos,
            'total_lancamentos': total_lancamentos,
            'extratos_conciliados': extratos_conciliados,
            'lancamentos_conciliados': lancamentos_conciliados,
            'percentual_extratos': (extratos_conciliados / total_extratos * 100) if total_extratos > 0 else 0,
            'percentual_lancamentos': (lancamentos_conciliados / total_lancamentos * 100) if total_lancamentos > 0 else 0
        },
        'categorias_extratos': categorias_extratos,
        'categorias_lancamentos': categorias_lancamentos
    })

@app.route('/api/upload-extrato', methods=['POST'])
@login_required
def upload_extrato():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Determinar formato do arquivo
        formato = 'CSV'
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            formato = 'EXCEL'
        elif filename.endswith('.ofx'):
            formato = 'OFX'
        elif filename.endswith('.cnab') or filename.endswith('.ret'):
            formato = 'CNAB'
        elif filename.endswith('.pdf'):
            formato = 'PDF'
        
        # Processar arquivo baseado no formato
        registros = []
        if formato == 'CSV':
            registros = process_csv_file(filepath)
        elif formato == 'EXCEL':
            registros = process_excel_file(filepath)
        elif formato == 'OFX':
            registros = process_ofx_file(filepath)
        elif formato == 'CNAB':
            registros = process_cnab_file(filepath)
        elif formato == 'PDF':
            registros = process_pdf_file(filepath)
        
        # Aplicar regras de conciliação e detectar duplicatas
        registros_importados = 0
        duplicatas_detectadas = 0
        
        for registro in registros:
            # Gerar hash da transação
            hash_transacao = generate_hash(registro['descricao'], registro['valor'], registro['data'], registro['tipo'])
            
            # Verificar se já existe
            if detect_duplicate_transaction(registro['descricao'], registro['valor'], registro['data'], registro['tipo'], 'extrato'):
                duplicatas_detectadas += 1
                continue
            
            # Aplicar regras de conciliação
            aplicar_regras_conciliacao(registro, 'extrato')
            
            # Criar registro
            extrato = ExtratoBancario(
                data=registro['data'],
                descricao=registro['descricao'],
                valor=registro['valor'],
                tipo=registro['tipo'],
                categoria=registro.get('categoria', 'Não categorizado'),
                arquivo_origem=filename,
                formato_arquivo=formato,
                numero_documento=registro.get('numero_documento', ''),
                hash_transacao=hash_transacao
            )
            
            db.session.add(extrato)
            registros_importados += 1
        
        db.session.commit()
        
        # Registrar na auditoria
        log_auditoria(
            'upload_extrato',
            'extrato_bancario',
            None,
            None,
            {
                'arquivo': filename,
                'formato': formato,
                'registros_importados': registros_importados,
                'duplicatas_detectadas': duplicatas_detectadas
            }
        )
        
        return jsonify({
            'success': True,
            'message': f'{registros_importados} registros importados com sucesso. {duplicatas_detectadas} duplicatas detectadas.',
            'registros': registros_importados,
            'duplicatas': duplicatas_detectadas
        })
    
    except Exception as e:
        logging.error(f"Erro no upload de extrato: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-lancamentos', methods=['POST'])
@login_required
def upload_lancamentos():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Processar arquivo
        registros = []
        if filename.endswith('.csv'):
            registros = process_csv_file(filepath)
        else:
            registros = process_excel_file(filepath)
        
        # Aplicar regras de conciliação e detectar duplicatas
        registros_importados = 0
        duplicatas_detectadas = 0
        
        for registro in registros:
            # Gerar hash da transação
            hash_transacao = generate_hash(registro['descricao'], registro['valor'], registro['data'], registro['tipo'])
            
            # Verificar se já existe
            if detect_duplicate_transaction(registro['descricao'], registro['valor'], registro['data'], registro['tipo'], 'lancamento'):
                duplicatas_detectadas += 1
                continue
            
            # Aplicar regras de conciliação
            aplicar_regras_conciliacao(registro, 'lancamento')
            
            # Criar registro
            lancamento = LancamentoContabil(
                data=registro['data'],
                descricao=registro['descricao'],
                valor=registro['valor'],
                tipo=registro['tipo'],
                categoria=registro.get('categoria', 'Não categorizado'),
                arquivo_origem=filename,
                numero_documento=registro.get('numero_documento', ''),
                centro_custo=registro.get('centro_custo', ''),
                conta_contabil=registro.get('conta_contabil', ''),
                fornecedor_cliente=registro.get('fornecedor_cliente', ''),
                hash_transacao=hash_transacao
            )
            
            db.session.add(lancamento)
            registros_importados += 1
        
        db.session.commit()
        
        # Registrar na auditoria
        log_auditoria(
            'upload_lancamentos',
            'lancamento_contabil',
            None,
            None,
            {
                'arquivo': filename,
                'registros_importados': registros_importados,
                'duplicatas_detectadas': duplicatas_detectadas
            }
        )
        
        return jsonify({
            'success': True,
            'message': f'{registros_importados} lançamentos importados com sucesso. {duplicatas_detectadas} duplicatas detectadas.',
            'registros': registros_importados,
            'duplicatas': duplicatas_detectadas
        })
    
    except Exception as e:
        logging.error(f"Erro no upload de lançamentos: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/extratos')
@login_required
def get_extratos():
    try:
        extratos = ExtratoBancario.query.order_by(ExtratoBancario.data.desc()).all()
        return jsonify([{
            'id': e.id,
            'data': e.data.strftime('%Y-%m-%d'),
            'descricao': e.descricao,
            'valor': e.valor,
            'tipo': e.tipo,
            'categoria': e.categoria,
            'conciliado': e.conciliado,
            'arquivo_origem': e.arquivo_origem,
            'formato_arquivo': e.formato_arquivo,
            'numero_documento': e.numero_documento,
            'transacao_recorrente': e.transacao_recorrente
        } for e in extratos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/lancamentos')
@login_required
def get_lancamentos():
    try:
        lancamentos = LancamentoContabil.query.order_by(LancamentoContabil.data.desc()).all()
        return jsonify([{
            'id': l.id,
            'data': l.data.strftime('%Y-%m-%d'),
            'descricao': l.descricao,
            'valor': l.valor,
            'tipo': l.tipo,
            'categoria': l.categoria,
            'conciliado': l.conciliado,
            'arquivo_origem': l.arquivo_origem,
            'numero_documento': l.numero_documento,
            'centro_custo': l.centro_custo,
            'conta_contabil': l.conta_contabil,
            'fornecedor_cliente': l.fornecedor_cliente
        } for l in lancamentos])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/conciliar', methods=['POST'])
@login_required
def conciliar():
    try:
        data = request.get_json()
        extrato_id = data.get('extrato_id')
        lancamento_id = data.get('lancamento_id')
        observacoes = data.get('observacoes', '')
        
        # Verificar se já existe conciliação
        conciliacao_existente = Conciliacao.query.filter_by(
            extrato_id=extrato_id, 
            lancamento_id=lancamento_id
        ).first()
        
        if conciliacao_existente:
            return jsonify({'error': 'Conciliação já existe'}), 400
        
        # Criar nova conciliação
        conciliacao = Conciliacao(
            extrato_id=extrato_id,
            lancamento_id=lancamento_id,
            usuario_id=current_user.id,
            observacoes=observacoes,
            tipo_conciliacao='manual'
        )
        db.session.add(conciliacao)
        
        # Marcar como conciliado
        extrato = ExtratoBancario.query.get(extrato_id)
        lancamento = LancamentoContabil.query.get(lancamento_id)
        
        if extrato:
            extrato.conciliado = True
        if lancamento:
            lancamento.conciliado = True
        
        db.session.commit()
        
        # Registrar na auditoria
        log_auditoria(
            'conciliar_manual',
            'conciliacao',
            conciliacao.id,
            None,
            {
                'extrato_id': extrato_id,
                'lancamento_id': lancamento_id,
                'observacoes': observacoes
            }
        )
        
        return jsonify({'success': True, 'message': 'Conciliação realizada com sucesso'})
    
    except Exception as e:
        logging.error(f"Erro na conciliação: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/desconciliar/<int:conciliacao_id>', methods=['DELETE'])
@login_required
def desconciliar(conciliacao_id):
    try:
        conciliacao = Conciliacao.query.get(conciliacao_id)
        if not conciliacao:
            return jsonify({'error': 'Conciliação não encontrada'}), 404
        
        # Desmarcar como conciliado
        if conciliacao.extrato:
            conciliacao.extrato.conciliado = False
        if conciliacao.lancamento:
            conciliacao.lancamento.conciliado = False
        
        # Registrar na auditoria antes de deletar
        log_auditoria(
            'desconciliar',
            'conciliacao',
            conciliacao_id,
            {
                'extrato_id': conciliacao.extrato_id,
                'lancamento_id': conciliacao.lancamento_id,
                'observacoes': conciliacao.observacoes
            },
            None
        )
        
        db.session.delete(conciliacao)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Conciliação removida com sucesso'})
    
    except Exception as e:
        logging.error(f"Erro ao desconciliar: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/estatisticas')
@login_required
def get_estatisticas():
    try:
        # Parâmetros de filtro de período
        periodo = request.args.get('periodo', 'total')  # hoje, semana, mes, ano, personalizado, total
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Definir datas baseadas no período
        hoje = datetime.now().date()
        
        if periodo == 'hoje':
            data_inicio = hoje
            data_fim = hoje
        elif periodo == 'semana':
            # Início da semana (segunda-feira)
            inicio_semana = hoje - timedelta(days=hoje.weekday())
            data_inicio = inicio_semana
            data_fim = hoje
        elif periodo == 'mes':
            # Início do mês
            data_inicio = hoje.replace(day=1)
            data_fim = hoje
        elif periodo == 'ano':
            # Início do ano
            data_inicio = hoje.replace(month=1, day=1)
            data_fim = hoje
        elif periodo == 'personalizado':
            if data_inicio:
                data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            if data_fim:
                data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            # Total (sem filtro de data)
            data_inicio = None
            data_fim = None
        
        # Queries base
        query_extratos = ExtratoBancario.query
        query_lancamentos = LancamentoContabil.query
        query_conciliacoes = Conciliacao.query
        query_divergencias = Divergencia.query
        
        # Aplicar filtros de data
        if data_inicio and data_fim:
            query_extratos = query_extratos.filter(
                ExtratoBancario.data >= data_inicio,
                ExtratoBancario.data <= data_fim
            )
            query_lancamentos = query_lancamentos.filter(
                LancamentoContabil.data >= data_inicio,
                LancamentoContabil.data <= data_fim
            )
            query_conciliacoes = query_conciliacoes.filter(
                Conciliacao.data_conciliacao >= datetime.combine(data_inicio, datetime.min.time()),
                Conciliacao.data_conciliacao <= datetime.combine(data_fim, datetime.max.time())
            )
            query_divergencias = query_divergencias.filter(
                Divergencia.created_at >= datetime.combine(data_inicio, datetime.min.time()),
                Divergencia.created_at <= datetime.combine(data_fim, datetime.max.time())
            )
        
        # Calcular estatísticas
        total_extratos = query_extratos.count()
        total_lancamentos = query_lancamentos.count()
        extratos_conciliados = query_extratos.filter_by(conciliado=True).count()
        lancamentos_conciliados = query_lancamentos.filter_by(conciliado=True).count()
        divergencias_pendentes = query_divergencias.filter_by(status='pendente').count()
        transacoes_recorrentes = query_extratos.filter_by(transacao_recorrente=True).count()
        total_conciliacoes = query_conciliacoes.count()
        
        # Calcular valores financeiros
        valor_total_extratos = query_extratos.with_entities(db.func.sum(ExtratoBancario.valor)).scalar() or 0
        valor_extratos_conciliados = query_extratos.filter_by(conciliado=True).with_entities(db.func.sum(ExtratoBancario.valor)).scalar() or 0
        
        return jsonify({
            'periodo': periodo,
            'data_inicio': data_inicio.strftime('%Y-%m-%d') if data_inicio else None,
            'data_fim': data_fim.strftime('%Y-%m-%d') if data_fim else None,
            'total_extratos': total_extratos,
            'total_lancamentos': total_lancamentos,
            'extratos_conciliados': extratos_conciliados,
            'lancamentos_conciliados': lancamentos_conciliados,
            'divergencias_pendentes': divergencias_pendentes,
            'transacoes_recorrentes': transacoes_recorrentes,
            'total_conciliacoes': total_conciliacoes,
            'valor_total_extratos': float(valor_total_extratos),
            'valor_extratos_conciliados': float(valor_extratos_conciliados),
            'percentual_extratos': (extratos_conciliados / total_extratos * 100) if total_extratos > 0 else 0,
            'percentual_lancamentos': (lancamentos_conciliados / total_lancamentos * 100) if total_lancamentos > 0 else 0,
            'percentual_conciliacao': (valor_extratos_conciliados / valor_total_extratos * 100) if valor_total_extratos > 0 else 0
        })
        
    except Exception as e:
        logging.error(f"Erro ao buscar estatísticas: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usuario')
def get_usuario():
    try:
        # Simulação de dados do usuário (em produção, usar dados reais do banco)
        if current_user.is_authenticated:
            # Usar o nome completo do usuário e calcular iniciais corretamente
            nome_completo = current_user.nome_completo if hasattr(current_user, 'nome_completo') and current_user.nome_completo else current_user.username
            
            # Calcular iniciais baseadas no nome completo
            palavras = nome_completo.split()
            iniciais = ''.join([palavra[0].upper() for palavra in palavras[:2]])
            
            # Determinar o perfil do usuário
            perfil_display = 'Administrador' if current_user.perfil == 'admin' else 'Auditor' if current_user.perfil == 'auditor' else 'Usuário'
            
            usuario = {
                'nome': nome_completo,
                'email': current_user.email if hasattr(current_user, 'email') else 'usuario@financesync.com',
                'iniciais': iniciais,
                'role': perfil_display,
                'ultimo_acesso': current_user.ultimo_acesso.strftime('%d/%m/%Y %H:%M') if hasattr(current_user, 'ultimo_acesso') and current_user.ultimo_acesso else 'Nunca',
                'avatar_url': None  # Pode ser implementado depois
            }
        else:
            # Dados padrão para usuário não logado
            usuario = {
                'nome': 'Leonardo Carlos',
                'email': 'leonardo@financesync.com', 
                'iniciais': 'LC',
                'role': 'Administrador',
                'ultimo_acesso': datetime.now().strftime('%d/%m/%Y %H:%M'),
                'avatar_url': None
            }
        
        return jsonify(usuario)
    except Exception as e:
        # Retornar dados padrão em caso de erro
        return jsonify({
            'nome': 'Usuário Sistema',
            'email': 'usuario@financesync.com',
            'iniciais': 'US',
            'role': 'Usuário',
            'ultimo_acesso': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'avatar_url': None
        })

@app.route('/api/limpar-dados', methods=['DELETE'])
@login_required
@admin_required
def limpar_dados():
    try:
        # Registrar na auditoria
        log_auditoria(
            'limpar_dados',
            'sistema',
            None,
            None,
            {'acao': 'Limpeza completa de dados'}
        )
        
        # Limpar todas as tabelas
        Conciliacao.query.delete()
        ExtratoBancario.query.delete()
        LancamentoContabil.query.delete()
        Divergencia.query.delete()
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Todos os dados foram removidos'})
    except Exception as e:
        logging.error(f"Erro ao limpar dados: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/criar-dados-exemplo', methods=['POST'])
@login_required
def criar_dados_exemplo():
    try:
        from random import uniform, choice, randint
        
        # Criar conta bancária exemplo se não existir
        conta = ContaBancaria.query.first()
        if not conta:
            conta = ContaBancaria(
                banco='Banco Exemplo',
                agencia='1234',
                conta='56789-0',
                tipo_conta='corrente',
                saldo_atual=10000.0
            )
            db.session.add(conta)
            db.session.commit()
        
        # Criar extratos de exemplo para diferentes períodos
        hoje = datetime.now().date()
        
        # Dados de hoje
        for i in range(5):
            extrato = ExtratoBancario(
                conta_id=conta.id,
                data=hoje,
                descricao=f'Transação do dia {i+1}',
                valor=round(uniform(100, 1000), 2),
                tipo=choice(['credito', 'debito']),
                categoria='Exemplo',
                arquivo_origem='dados_exemplo.csv'
            )
            db.session.add(extrato)
        
        # Dados desta semana
        for i in range(10):
            data_semana = hoje - timedelta(days=randint(1, 7))
            extrato = ExtratoBancario(
                conta_id=conta.id,
                data=data_semana,
                descricao=f'Transação da semana {i+1}',
                valor=round(uniform(50, 800), 2),
                tipo=choice(['credito', 'debito']),
                categoria='Exemplo',
                arquivo_origem='dados_exemplo.csv'
            )
            db.session.add(extrato)
        
        # Dados deste mês
        for i in range(20):
            data_mes = hoje - timedelta(days=randint(8, 30))
            extrato = ExtratoBancario(
                conta_id=conta.id,
                data=data_mes,
                descricao=f'Transação do mês {i+1}',
                valor=round(uniform(25, 1500), 2),
                tipo=choice(['credito', 'debito']),
                categoria='Exemplo',
                arquivo_origem='dados_exemplo.csv'
            )
            db.session.add(extrato)
        
        # Lançamentos contábeis correspondentes
        extratos = ExtratoBancario.query.filter_by(arquivo_origem='dados_exemplo.csv').all()
        for extrato in extratos[:15]:  # Conciliar apenas alguns
            lancamento = LancamentoContabil(
                data=extrato.data,
                descricao=extrato.descricao,
                valor=extrato.valor,
                tipo=extrato.tipo,
                categoria=extrato.categoria,
                arquivo_origem='dados_exemplo.csv'
            )
            db.session.add(lancamento)
            
            # Marcar como conciliado
            extrato.conciliado = True
            lancamento.conciliado = True
            
            # Criar registro de conciliação
            conciliacao = Conciliacao(
                extrato_id=extrato.id,
                lancamento_id=lancamento.id,
                usuario_id=current_user.id if current_user.is_authenticated else 1,
                tipo_conciliacao='automatica',
                status='ativa'
            )
            db.session.add(conciliacao)
        
        # Criar algumas divergências
        for i in range(3):
            divergencia = Divergencia(
                tipo='valor_incorreto',
                descricao=f'Divergência de exemplo {i+1}',
                status='pendente'
            )
            db.session.add(divergencia)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Dados de exemplo criados com sucesso!',
            'extratos_criados': len(extratos),
            'conciliacoes_criadas': 15,
            'divergencias_criadas': 3
        })
        
    except Exception as e:
        logging.error(f"Erro ao criar dados de exemplo: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/limpar-todos-dados', methods=['POST'])
@login_required
def limpar_todos_dados():
    try:
        # Deletar todos os dados em ordem (respeitando foreign keys)
        Conciliacao.query.delete()
        Divergencia.query.delete()
        ExtratoBancario.query.delete()
        LancamentoContabil.query.delete()
        ContaBancaria.query.delete()
        RegraConciliacao.query.delete()
        
        db.session.commit()
        
        logging.info("Todos os dados foram limpos do sistema")
        
        return jsonify({
            'success': True,
            'message': 'Todos os dados foram removidos com sucesso!'
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao limpar dados: {e}")
        return jsonify({'error': str(e)}), 500

# === ROTAS DE ADMINISTRAÇÃO ===

@app.route('/api/admin/usuarios', methods=['GET'])
@login_required
@admin_required
def get_admin_usuarios():
    """Listar todos os usuários"""
    try:
        usuarios = Usuario.query.all()
        
        resultado = [{
            'id': u.id,
            'username': u.username,
            'email': u.email,
            'nome_completo': u.nome_completo,
            'perfil': u.perfil,
            'ativo': u.ativo,
            'created_at': u.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'ultimo_acesso': u.ultimo_acesso.strftime('%Y-%m-%d %H:%M:%S') if u.ultimo_acesso else 'Nunca',
            'estatisticas': {
                'extratos': ExtratoBancario.query.filter_by(usuario_id=u.id).count(),
                'lancamentos': LancamentoContabil.query.filter_by(usuario_id=u.id).count(),
                'conciliacoes': Conciliacao.query.filter_by(usuario_id=u.id).count()
            }
        } for u in usuarios]
        
        return jsonify({'usuarios': resultado})
        
    except Exception as e:
        logging.error(f"Erro ao buscar usuários: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:usuario_id>/perfil', methods=['PUT'])
@login_required
@admin_required
def alterar_perfil_usuario(usuario_id):
    """Alterar perfil/permissões de um usuário"""
    try:
        data = request.get_json()
        novo_perfil = data.get('perfil')
        
        if novo_perfil not in ['admin', 'auditor', 'usuario']:
            return jsonify({'error': 'Perfil inválido'}), 400
        
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        # Não permitir alterar o próprio perfil
        if usuario.id == current_user.id:
            return jsonify({'error': 'Não é possível alterar o próprio perfil'}), 403
        
        perfil_anterior = usuario.perfil
        usuario.perfil = novo_perfil
        db.session.commit()
        
        # Log da alteração
        log_auditoria(
            'alterar_perfil_usuario',
            'usuario',
            usuario.id,
            {'perfil': perfil_anterior},
            {'perfil': novo_perfil}
        )
        
        return jsonify({
            'success': True,
            'message': f'Perfil do usuário {usuario.username} alterado de {perfil_anterior} para {novo_perfil}'
        })
        
    except Exception as e:
        logging.error(f"Erro ao alterar perfil do usuário: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:usuario_id>/status', methods=['PUT'])
@login_required
@admin_required
def alterar_status_usuario(usuario_id):
    """Ativar/desativar um usuário"""
    try:
        data = request.get_json()
        novo_status = data.get('ativo')
        
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        # Não permitir desativar o próprio usuário
        if usuario.id == current_user.id:
            return jsonify({'error': 'Não é possível alterar o próprio status'}), 403
        
        status_anterior = usuario.ativo
        usuario.ativo = novo_status
        db.session.commit()
        
        # Log da alteração
        log_auditoria(
            'alterar_status_usuario',
            'usuario',
            usuario.id,
            {'ativo': status_anterior},
            {'ativo': novo_status}
        )
        
        return jsonify({
            'success': True,
            'message': f'Usuário {usuario.username} {"ativado" if novo_status else "desativado"} com sucesso'
        })
        
    except Exception as e:
        logging.error(f"Erro ao alterar status do usuário: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:usuario_id>/resetar-senha', methods=['POST'])
@login_required
@admin_required
def resetar_senha_usuario(usuario_id):
    """Resetar senha de um usuário"""
    try:
        data = request.get_json()
        nova_senha = data.get('nova_senha', 'senha123')
        
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({'error': 'Usuário não encontrado'}), 404
        
        usuario.password_hash = generate_password_hash(nova_senha)
        db.session.commit()
        
        # Log da alteração
        log_auditoria(
            'resetar_senha_usuario',
            'usuario',
            usuario.id,
            None,
            {'acao': 'senha_resetada'}
        )
        
        return jsonify({
            'success': True,
            'message': f'Senha do usuário {usuario.username} resetada com sucesso',
            'nova_senha': nova_senha
        })
        
    except Exception as e:
        logging.error(f"Erro ao resetar senha do usuário: {e}")
        return jsonify({'error': str(e)}), 500

# === ROTAS DO PERFIL DO USUÁRIO ===

@app.route('/api/usuario/perfil', methods=['GET'])
@login_required
def get_perfil_usuario():
    """Obter dados do perfil do usuário atual"""
    try:
        usuario = current_user
        
        # Calcular estatísticas do usuário
        estatisticas = {
            'extratos': ExtratoBancario.query.filter_by(usuario_id=usuario.id).count(),
            'lancamentos': LancamentoContabil.query.filter_by(usuario_id=usuario.id).count(),
            'conciliacoes': Conciliacao.query.filter_by(usuario_id=usuario.id).count(),
            'divergencias': Conciliacao.query.filter_by(usuario_id=usuario.id, status='divergente').count()
        }
        
        return jsonify({
            'success': True,
            'usuario': {
                'id': usuario.id,
                'username': usuario.username,
                'email': usuario.email,
                'nome_completo': usuario.nome_completo,
                'perfil': usuario.perfil,
                'ativo': usuario.ativo,
                'created_at': usuario.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'ultimo_acesso': usuario.ultimo_acesso.strftime('%Y-%m-%d %H:%M:%S') if usuario.ultimo_acesso else 'Nunca',
                'foto_perfil': getattr(usuario, 'foto_perfil', None),
                'telefone': getattr(usuario, 'telefone', ''),
                'estatisticas': estatisticas
            }
        })
        
    except Exception as e:
        logging.error(f"Erro ao buscar perfil do usuário: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usuario/perfil', methods=['PUT'])
@login_required
def atualizar_perfil_usuario():
    """Atualizar dados do perfil do usuário atual"""
    try:
        data = request.get_json()
        usuario = current_user
        
        # Atualizar campos permitidos
        if 'nome_completo' in data:
            usuario.nome_completo = data['nome_completo']
        
        if 'email' in data:
            # Verificar se o email já está em uso por outro usuário
            existing_user = Usuario.query.filter(
                Usuario.email == data['email'], 
                Usuario.id != usuario.id
            ).first()
            
            if existing_user:
                return jsonify({'error': 'Este email já está em uso por outro usuário'}), 400
            
            usuario.email = data['email']
        
        if 'telefone' in data:
            # Adicionar campo telefone se não existir na tabela
            if hasattr(usuario, 'telefone'):
                usuario.telefone = data['telefone']
        
        db.session.commit()
        log_auditoria('atualizar_perfil', 'usuario', usuario.id, None, data)
        
        return jsonify({
            'success': True,
            'message': 'Perfil atualizado com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao atualizar perfil do usuário: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usuario/alterar-senha', methods=['POST'])
@login_required
def alterar_senha_usuario():
    """Alterar senha do usuário atual"""
    try:
        data = request.get_json()
        usuario = current_user
        
        senha_atual = data.get('senha_atual')
        nova_senha = data.get('nova_senha')
        
        if not senha_atual or not nova_senha:
            return jsonify({'error': 'Senha atual e nova senha são obrigatórias'}), 400
        
        # Verificar senha atual
        if not check_password_hash(usuario.password_hash, senha_atual):
            return jsonify({'error': 'Senha atual incorreta'}), 400
        
        # Validar nova senha
        if len(nova_senha) < 6:
            return jsonify({'error': 'A nova senha deve ter pelo menos 6 caracteres'}), 400
        
        # Atualizar senha
        usuario.password_hash = generate_password_hash(nova_senha)
        db.session.commit()
        
        log_auditoria('alterar_senha', 'usuario', usuario.id, None, {
            'usuario_id': usuario.id,
            'usuario': usuario.username
        })
        
        return jsonify({
            'success': True,
            'message': 'Senha alterada com sucesso'
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Erro ao alterar senha do usuário: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 